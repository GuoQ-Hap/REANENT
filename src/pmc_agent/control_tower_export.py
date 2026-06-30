from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pmc_agent.connectors.database import StiDatabaseConnector
from pmc_agent.schema_catalog import ALL_WAREHOUSE_CATALOG


PICI_HORIZONS = (7, 14, 21, 28, 35, 42, 49, 56, 63, 70, 77, 84, 98)
OVERSTOCK_AVAILABLE_DAY_RULES: dict[int, dict[str, Any]] = {
    1: {
        "label": "可售天数1",
        "thresholds": {"boom_wang": 90, "flat_stagnant": 60},
        "action": "重点监控运营清货进度",
    },
    2: {
        "label": "可售天数2",
        "thresholds": {"boom_wang": 120, "flat_stagnant": 105},
        "action": "禁止向FBA补货",
    },
    3: {
        "label": "可售天数3",
        "thresholds": {"boom_wang": 120, "flat_stagnant": 105},
        "action": "禁止向FBA补货",
    },
    4: {
        "label": "可售天数4",
        "thresholds": {"boom_wang": 120, "flat_stagnant": 105},
        "action": "禁止向海外仓补货；禁止向FBA补货",
    },
    5: {
        "label": "可售天数5",
        "thresholds": {"boom_wang": 180, "flat_stagnant": 150},
        "action": "禁止本地仓补货",
    },
    6: {
        "label": "可售天数6",
        "thresholds": {"boom_wang": 180, "flat_stagnant": 150},
        "action": "禁止本地仓补货；停止下采购单",
    },
}

SHEET_NAMES = [
    "sheet1",
    "W20-FNSKU断货分析",
    "12.28V1预测",
    "12.27V0",
    "实发差异",
    "实销",
    "周度超卖",
]

SHEET1_HEADERS = [
    "店铺MSKUFNSKU",
    "MSKU",
    "SKU",
    "FNSKU",
    "ASIN",
    "父ASIN",
    "品牌",
    "店铺",
    "销售部门",
    "销售员",
    "产品经理",
    "国家",
    "渠道",
    "品名",
    "seller_id",
    "发货国家",
    "区域",
    "colour",
    "色系",
    "size",
    "3月12V1MSKU销售属性",
    "MSKU销售属性",
    "MSKU产品属性",
    "是否为引流MSKU",
    "季节属性",
    "MSKU销售状态",
    "MSKU生命周期",
    "￥销售单价",
    "$销售单价",
    "物流模式",
    "头程物流渠道",
    "下单时长",
    "生产时长",
    "本地仓出库时长",
    "海外仓出库时长",
    "本地仓到FBA仓时长",
    "本地仓到海外仓时长",
    "海外仓到FBA仓时长",
    "可超卖比例",
    "销售安全库存",
    "销售安全库存天数",
    "物流安全库存天数",
    "生产安全库存天数",
    "FBA安全天数_fnsku",
    "海外仓安全天数_fnsku",
    "本地仓安全天数_fnsku",
    "FBA交付时长_fnsku",
    "海外仓交付时长_fnsku",
    "本地仓交付时长_fnsku",
    "FBA可售",
    "待调仓",
    "调仓中",
    "入库中",
    "实际在途",
    "计划入库",
    "发货计划量",
    "海外仓待检待上架",
    "海外仓调拨在途",
    "海外仓计划量",
    "本地仓待检待上架",
    "本地仓调拨在途",
    "待到货量",
    "采购计划量",
    "FBA仓_FNSKU",
    "海外仓_FNSKU",
    "海外仓在途_FNSKU",
    "本地仓_FNSKU",
    "本地仓在途_FNSKU",
    "90天毛利率",
    "库存1_FNSKU",
    "库存2_FNSKU",
    "库存3_FNSKU",
    "库存4_FNSKU",
    "库存5_FNSKU",
    "库存6_FNSKU",
    "可售天数1_FNSKU",
    "可售天数2_FNSKU",
    "可售天数3_FNSKU",
    "可售天数4_FNSKU",
    "可售天数5_FNSKU",
    "可售天数6_FNSKU",
    "可售天数1_str_FNSKU",
    "可售天数2_str_FNSKU",
    "可售天数3_str_FNSKU",
    "可售天数4_str_FNSKU",
    "可售天数5_str_FNSKU",
    "可售天数6_str_FNSKU",
    "补货倒计时1_FNSKU",
    "补货倒计时2_FNSKU",
    "补货倒计时3_FNSKU",
    "补货倒计时4_FNSKU",
    "补货倒计时5_FNSKU",
    "补货倒计时6_FNSKU",
    "fnsku_restocking_countdown_str_1",
    "fnsku_restocking_countdown_str_2",
    "fnsku_restocking_countdown_str_3",
    "fnsku_restocking_countdown_str_4",
    "fnsku_restocking_countdown_str_5",
    "fnsku_restocking_countdown_str_6",
    "断货风险1_FNSKU",
    "断货风险2_FNSKU",
    "断货风险3_FNSKU",
    "断货风险4_FNSKU",
    "断货风险5_FNSKU",
    "断货风险6_FNSKU",
    "FBA库龄1~30天",
    "FBA库龄31~60天",
    "FBA库龄61~90天",
    "FBA库龄91~180天",
    "FBA库龄181~270天",
    "FBA库龄271~330天",
    "FBA库龄331~365天",
    "FBA库龄365",
    "7天历史日均销量",
    "14天历史日均销量",
    "30天历史日均销量",
    "60天历史日均销量",
    "90天历史日均销量",
    "同期30天销量",
    "同期60天销量",
    "同期90天销量",
    "可售天数1_FNSKU_历史销量",
    "可售天数2_FNSKU_历史销量",
    "可售天数3_FNSKU_历史销量",
    "可售天数4_FNSKU_历史销量",
    "可售天数5_FNSKU_历史销量",
    "可售天数6_FNSKU_历史销量",
    "sale_fnsku_available_str_1",
    "sale_fnsku_available_str_2",
    "sale_fnsku_available_str_3",
    "sale_fnsku_available_str_4",
    "sale_fnsku_available_str_5",
    "sale_fnsku_available_str_6",
    "历史补货倒计时1_FNSKU",
    "历史补货倒计时2_FNSKU",
    "历史补货倒计时3_FNSKU",
    "历史补货倒计时4_FNSKU",
    "历史补货倒计时5_FNSKU",
    "历史补货倒计时6_FNSKU",
    "sale_fnsku_restocking_countdown_str_1",
    "sale_fnsku_restocking_countdown_str_2",
    "sale_fnsku_restocking_countdown_str_3",
    "sale_fnsku_restocking_countdown_str_4",
    "sale_fnsku_restocking_countdown_str_5",
    "sale_fnsku_restocking_countdown_str_6",
    "历史断货风险1_FNSKU",
    "历史断货风险2_FNSKU",
    "历史断货风险3_FNSKU",
    "历史断货风险4_FNSKU",
    "历史断货风险5_FNSKU",
    "历史断货风险6_FNSKU",
    "15天预估销量",
    "30天预估销量",
    "45天预估销量",
    "60天预估销量",
    "75天预估销量",
    "90天预估销量",
    "chazhi_0_7",
    "chazhi_0_14",
    "chazhi_0_21",
    "chazhi_0_28",
    "chazhi_0_35",
    "chazhi_0_42",
    "预计断货数量",
    "chazhi_0_49",
    "chazhi_0_56",
    "chazhi_0_63",
    "chazhi_0_70",
    "chazhi_0_77",
    "chazhi_0_84",
    "chazhi_0_98",
    "2026W19",
    "2026W20",
    "2026W21",
    "2026W22",
    "2026W23",
    "2026W24",
    "2026W25",
    "2026W26",
    "2026W27",
    "2026W28",
    "2026W29",
    "2026W30",
    "2026W31",
    "2026W32",
    "2026W33",
    "2026W34",
    "2026W35",
    "2026W36",
    "2026W37",
    "2026W38",
    "2026W39",
    "2026W40",
    "2026W41",
    "2026W42",
    "2026W43",
    "2026W44",
    "2026W45",
    "2026W46",
    "2026W47",
    "12.28-5.17总预估（取值12.28V1库存底表）",
    "12.28-5.17总销量",
    "总预实差异",
    "总体超卖比例",
    "总体售卖问题",
    "4/13V0版5月11~5月17预测销量",
    "12.28V1版5月11~5月17预测销量",
    "周度新旧预估差",
    "5月11~5月17实销",
    "超卖（对比12.28预估）",
    "超卖比例（对比12.28预估）",
    "当周售卖问题",
    "在途断货分析",
    "物流问题（匹配物流批次表）",
    "物流原因",
    "计划发货量",
    "运营确认发货量",
    "差异",
    "主要责任部门",
    "总体原因",
    "响应措施",
]


HEADER_FIELD_MAP = {
    "MSKU": "msku",
    "SKU": "sku",
    "FNSKU": "fnsku",
    "ASIN": "asin",
    "父ASIN": "parent_asin",
    "品牌": "brand",
    "店铺": "store_name",
    "销售部门": "sales_apartment",
    "销售员": "salesman",
    "产品经理": "product_manager",
    "国家": "country_code",
    "渠道": "channel",
    "品名": "sku_name",
    "seller_id": "seller_id",
    "发货国家": "shipments_country",
    "区域": "region",
    "colour": "colour",
    "色系": "color",
    "size": "size",
    "MSKU销售属性": "msku_sales_property",
    "MSKU产品属性": "msku_product_property",
    "是否为引流MSKU": "msku_isattract",
    "季节属性": "seasonality",
    "MSKU销售状态": "msku_status",
    "MSKU生命周期": "msku_life_process",
    "￥销售单价": "last_30_order_price",
    "$销售单价": "last_30_order_us_price",
    "物流模式": "logistics_model",
    "头程物流渠道": "first_leg_logistics_channel",
    "下单时长": "order_duration",
    "生产时长": "production_duration",
    "本地仓出库时长": "local_warehouse_pick_time",
    "海外仓出库时长": "overseas_warehouse_pick_time",
    "本地仓到FBA仓时长": "local_to_FBA_time",
    "本地仓到海外仓时长": "local_to_overseas_warehouse_time",
    "海外仓到FBA仓时长": "overseas_to_FBA_time",
    "可超卖比例": "max_over_sell_ratio",
    "销售安全库存": "safety_stock_sales",
    "销售安全库存天数": "safety_stock_days_sales",
    "物流安全库存天数": "logistics_safety_stock_days",
    "生产安全库存天数": "production_safety_stock_days",
    "FBA安全天数_fnsku": "fba_safety_days_fn",
    "海外仓安全天数_fnsku": "overseas_warehouse_safety_days",
    "本地仓安全天数_fnsku": "local_warehouse_safety_days",
    "FBA交付时长_fnsku": "FBA_delivery_time_fn",
    "海外仓交付时长_fnsku": "overseas_warehouse_delivery_time_fn",
    "本地仓交付时长_fnsku": "local_warehouse_delivery_time_fn",
    "FBA可售": "afn_fulfillable_quantity",
    "待调仓": "reserved_fc_transfers",
    "调仓中": "reserved_fc_processing",
    "入库中": "afn_inbound_receiving_quantity",
    "实际在途": "stock_up_num",
    "计划入库": "afn_inbound_working_quantity",
    "海外仓待检待上架": "overseas_wh_product_qc_num",
    "海外仓调拨在途": "overseas_wh_product_onway",
    "海外仓计划量": "planned_quantity",
    "本地仓待检待上架": "local_wh_product_qc_num",
    "本地仓调拨在途": "local_wh_product_onway",
    "待到货量": "quantity_receive",
    "采购计划量": "procurement_plan_quantity",
    "FBA仓_FNSKU": "fba_warehouse_quantity",
    "海外仓_FNSKU": "overseas_warehouse_quantity",
    "海外仓在途_FNSKU": "oversease_afn_inbound_shipped_quantity",
    "本地仓_FNSKU": "local_warehouse_quantity",
    "本地仓在途_FNSKU": "local_afn_inbound_shipped_quantity",
    "90天毛利率": "last_90_gross_margin",
    "库存1_FNSKU": "fnsku_inventory_1",
    "库存2_FNSKU": "fnsku_inventory_2",
    "库存3_FNSKU": "fnsku_inventory_3",
    "库存4_FNSKU": "fnsku_inventory_4",
    "库存5_FNSKU": "fnsku_inventory_5",
    "库存6_FNSKU": "fnsku_inventory_6",
    "可售天数1_FNSKU": "fnsku_available_days_1",
    "可售天数2_FNSKU": "fnsku_available_days_2",
    "可售天数3_FNSKU": "fnsku_available_days_3",
    "可售天数4_FNSKU": "fnsku_available_days_4",
    "可售天数5_FNSKU": "fnsku_available_days_5",
    "可售天数6_FNSKU": "fnsku_available_days_6",
    "可售天数1_str_FNSKU": "fnsku_available_days_str_1",
    "可售天数2_str_FNSKU": "fnsku_available_days_str_2",
    "可售天数3_str_FNSKU": "fnsku_available_days_str_3",
    "可售天数4_str_FNSKU": "fnsku_available_days_str_4",
    "可售天数5_str_FNSKU": "fnsku_available_days_str_5",
    "可售天数6_str_FNSKU": "fnsku_available_days_str_6",
    "补货倒计时1_FNSKU": "fnsku_restocking_countdown_1",
    "补货倒计时2_FNSKU": "fnsku_restocking_countdown_2",
    "补货倒计时3_FNSKU": "fnsku_restocking_countdown_3",
    "补货倒计时4_FNSKU": "fnsku_restocking_countdown_4",
    "补货倒计时5_FNSKU": "fnsku_restocking_countdown_5",
    "补货倒计时6_FNSKU": "fnsku_restocking_countdown_6",
    "断货风险1_FNSKU": "fnsku_out_of_stock_risk_1",
    "断货风险2_FNSKU": "fnsku_out_of_stock_risk_2",
    "断货风险3_FNSKU": "fnsku_out_of_stock_risk_3",
    "断货风险4_FNSKU": "fnsku_out_of_stock_risk_4",
    "断货风险5_FNSKU": "fnsku_out_of_stock_risk_5",
    "断货风险6_FNSKU": "fnsku_out_of_stock_risk_6",
    "FBA库龄1~30天": "inv_age_0_to_30_days",
    "FBA库龄31~60天": "inv_age_31_to_60_days",
    "FBA库龄61~90天": "inv_age_61_to_90_days",
    "FBA库龄91~180天": "inv_age_91_to_180_days",
    "FBA库龄181~270天": "inv_age_181_to_270_days",
    "FBA库龄271~330天": "inv_age_271_to_330_days",
    "FBA库龄331~365天": "inv_age_331_to_365_days",
    "FBA库龄365": "inv_age_365_plus_days",
    "7天历史日均销量": "sale_quantity_7",
    "14天历史日均销量": "sale_quantity_14",
    "30天历史日均销量": "sale_quantity_30",
    "60天历史日均销量": "sale_quantity_60",
    "90天历史日均销量": "sale_quantity_90",
    "同期30天销量": "last_year_sale_quantity_30",
    "同期60天销量": "last_year_sale_quantity_60",
    "同期90天销量": "last_year_sale_quantity_90",
    "可售天数1_FNSKU_历史销量": "sale_fnsku_available_1",
    "可售天数2_FNSKU_历史销量": "sale_fnsku_available_2",
    "可售天数3_FNSKU_历史销量": "sale_fnsku_available_3",
    "可售天数4_FNSKU_历史销量": "sale_fnsku_available_4",
    "可售天数5_FNSKU_历史销量": "sale_fnsku_available_5",
    "可售天数6_FNSKU_历史销量": "sale_fnsku_available_6",
    "15天预估销量": "future_15d_sales",
    "30天预估销量": "future_30d_sales",
    "45天预估销量": "future_45d_sales",
    "60天预估销量": "future_60d_sales",
    "75天预估销量": "future_75d_sales",
    "90天预估销量": "future_90d_sales",
}

for index in range(1, 7):
    HEADER_FIELD_MAP[f"sale_fnsku_available_str_{index}"] = f"sale_fnsku_available_str_{index}"
    HEADER_FIELD_MAP[f"历史补货倒计时{index}_FNSKU"] = f"sale_fnsku_restocking_countdown_{index}"
    HEADER_FIELD_MAP[f"sale_fnsku_restocking_countdown_str_{index}"] = f"sale_fnsku_restocking_countdown_str_{index}"
    HEADER_FIELD_MAP[f"历史断货风险{index}_FNSKU"] = f"sale_fnsku_out_of_stock_risk_{index}"
    HEADER_FIELD_MAP[f"fnsku_restocking_countdown_str_{index}"] = f"fnsku_restocking_countdown_str_{index}"

SOURCE_FIELDS = tuple(dict.fromkeys(field for field in HEADER_FIELD_MAP.values() if field))


def build_daily_investigation_workbook(
    filters: dict[str, Any] | None = None,
    max_rows: int = 20000,
    connector: StiDatabaseConnector | None = None,
) -> tuple[bytes, dict[str, Any]]:
    connector = connector or StiDatabaseConnector()
    rows = connector.get_inventory_export_rows(SOURCE_FIELDS, filters=filters, limit=max_rows)
    pici_rows = connector.get_pici_shortage_rows(store_name=_selected_store(filters))
    pici_by_key = {
        (_norm(row.get("fnsku")), _text(row.get("store_name"))): row
        for row in pici_rows
        if _norm(row.get("fnsku"))
    }
    output_rows = []
    reason_counts = {"shortage": 0, "overstock": 0, "both": 0}
    for row in rows:
        pici = pici_by_key.get((_norm(row.get("fnsku")), _text(row.get("store_name"))), {})
        enriched = {**row, **{f"chazhi_0_{horizon}": pici.get(f"chazhi_0_{horizon}") for horizon in PICI_HORIZONS}}
        reason = _investigation_reason(enriched)
        if not reason["include"]:
            continue
        output_rows.append(_sheet1_row(enriched, reason))
        if reason["shortage"] and reason["overstock"]:
            reason_counts["both"] += 1
        elif reason["shortage"]:
            reason_counts["shortage"] += 1
        elif reason["overstock"]:
            reason_counts["overstock"] += 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAMES[0]
    _write_sheet1(sheet, output_rows)
    for name in SHEET_NAMES[1:]:
        ws = workbook.create_sheet(name)
        ws.append(["待接入", "说明"])
        ws.append([name, "第一步先生成 sheet1，后续按参考文件逐张补齐。"])
        ws.freeze_panes = "A2"
    stream = BytesIO()
    workbook.save(stream)
    meta = {
        "source_rows": len(rows),
        "export_rows": len(output_rows),
        "reason_counts": reason_counts,
        "sheet_names": SHEET_NAMES,
    }
    return stream.getvalue(), meta


def export_filename() -> str:
    return f"爆旺断货冗余排查_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


def _write_sheet1(sheet: Any, rows: list[list[Any]]) -> None:
    sheet.append(SHEET1_HEADERS)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(SHEET1_HEADERS))}{max(1, sheet.max_row)}"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, len(SHEET1_HEADERS) + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 16
    for column in (1, 14, 213, 219, 220, 221):
        sheet.column_dimensions[get_column_letter(column)].width = 28


def _sheet1_row(row: dict[str, Any], reason: dict[str, Any]) -> list[Any]:
    values = []
    for header in SHEET1_HEADERS:
        if header == "店铺MSKUFNSKU":
            values.append(f"{_text(row.get('store_name'))}{_text(row.get('msku'))}{_text(row.get('fnsku'))}")
        elif header.startswith("chazhi_"):
            values.append(row.get(header))
        elif header == "预计断货数量":
            values.append(reason["shortage_qty"])
        elif header == "在途断货分析":
            values.append("断货排查" if reason["shortage"] else "")
        elif header == "总体售卖问题":
            values.append(reason["sell_issue"])
        elif header == "当周售卖问题":
            values.append(reason["sell_issue"])
        elif header == "主要责任部门":
            values.append(reason["owner"])
        elif header == "总体原因":
            values.append(reason["reason_text"])
        elif header == "响应措施":
            values.append(reason["action"])
        else:
            values.append(row.get(HEADER_FIELD_MAP.get(header, ""), ""))
    return values


def _investigation_reason(row: dict[str, Any]) -> dict[str, Any]:
    sales_property = _text(row.get("msku_sales_property"))
    chazhi_inventory = _parse_chazhi_inventory(_text(row.get("chazhi_0_7")))
    shortage_threshold = 4 if sales_property in {"爆", "旺"} else 1
    shortage = chazhi_inventory is not None and chazhi_inventory <= shortage_threshold
    overstock_hits = _overstock_available_day_hits(row, sales_property)
    overstock = bool(overstock_hits)
    shortage_qty = _min_chazhi_gap(row)
    dominant_hit = (
        max(overstock_hits, key=lambda item: item["days"] - item["threshold"])
        if overstock_hits
        else None
    )
    inventory = _number(row.get(f"fnsku_inventory_{dominant_hit['index']}")) if dominant_hit else 0
    if not inventory and overstock:
        inventory = _chain_inventory(row)
    forecast_90 = _number(row.get("future_90d_sales"))
    threshold_days = dominant_hit["threshold"] if dominant_hit else 0
    forecast_threshold = forecast_90 * threshold_days / 90 if forecast_90 and threshold_days else 0
    redundant_qty = max(0, round(inventory - forecast_threshold, 2))
    reasons = []
    actions = []
    if shortage:
        reasons.append(
            f"chazhi_0_7库存量={chazhi_inventory}，{sales_property or '未知属性'}阈值<={shortage_threshold}"
        )
        actions.append("断货排查：核查库存、在途和补货计划")
    if overstock:
        hit_text = "；".join(
            f"{hit['label']}({hit['group_label']})={hit['days']:g}>{hit['threshold']:g}"
            for hit in overstock_hits
        )
        hit_actions = _unique_text([str(hit["action"]) for hit in overstock_hits])
        reasons.append(f"{hit_text}，预计冗余量={redundant_qty}")
        actions.append("冗余排查：冻结SKU、复核发货计划、" + "；".join(hit_actions))
    if shortage and overstock:
        sell_issue = "断货+冗余"
        owner = "PMC/运营"
    elif shortage:
        sell_issue = "断货"
        owner = "PMC/物流/运营"
    elif overstock:
        sell_issue = "冗余"
        owner = "PMC/运营"
    else:
        sell_issue = ""
        owner = ""
    return {
        "include": shortage or overstock,
        "shortage": shortage,
        "overstock": overstock,
        "shortage_qty": shortage_qty,
        "sell_issue": sell_issue,
        "owner": owner,
        "reason_text": "；".join(reasons),
        "action": "；".join(actions),
    }


def _overstock_available_day_hits(row: dict[str, Any], sales_property: str) -> list[dict[str, Any]]:
    group = _sales_property_threshold_group(sales_property)
    group_label = "平滞" if group == "flat_stagnant" else "爆旺"
    hits = []
    for index, rule in OVERSTOCK_AVAILABLE_DAY_RULES.items():
        days = _number(row.get(f"fnsku_available_days_{index}"))
        threshold = rule["thresholds"][group]
        if days > threshold:
            hits.append(
                {
                    "index": index,
                    "label": rule["label"],
                    "group_label": group_label,
                    "days": days,
                    "threshold": threshold,
                    "action": rule["action"],
                }
            )
    return hits


def _sales_property_threshold_group(sales_property: str) -> str:
    text = str(sales_property or "").strip()
    return "flat_stagnant" if "平" in text or "滞" in text else "boom_wang"


def _unique_text(values: list[str]) -> list[str]:
    unique = []
    for value in values:
        text = value.strip()
        if text and text not in unique:
            unique.append(text)
    return unique


def _chain_inventory(row: dict[str, Any]) -> float:
    return sum(
        _number(row.get(field))
        for field in (
            "fba_warehouse_quantity",
            "overseas_warehouse_quantity",
            "oversease_afn_inbound_shipped_quantity",
            "local_warehouse_quantity",
            "local_afn_inbound_shipped_quantity",
        )
    )


def _parse_chazhi_inventory(value: str) -> float | None:
    if not value or "/" not in value:
        return None
    try:
        return float(value.split("/", 1)[0])
    except ValueError:
        return None


def _min_chazhi_gap(row: dict[str, Any]) -> float | None:
    gaps = []
    for horizon in (7, 14, 21, 28, 35, 42):
        value = _text(row.get(f"chazhi_0_{horizon}"))
        if "(" not in value or ")" not in value:
            continue
        try:
            gaps.append(float(value.rsplit("(", 1)[-1].split(")", 1)[0]))
        except ValueError:
            continue
    return min(gaps) if gaps else None


def _selected_store(filters: dict[str, Any] | None) -> str | None:
    if not filters:
        return None
    value = filters.get("store_name")
    return _text(value).strip() or None


def _number(value: Any) -> float:
    if value is None:
        return 0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def _text(value: Any) -> str:
    return "" if value is None else str(value)


def _norm(value: Any) -> str:
    return _text(value).strip().upper()
