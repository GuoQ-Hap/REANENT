from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pmc_agent.connectors.database import StiDatabaseConnector
from pmc_agent.control_tower import ControlTowerItem, get_control_tower_summary


MAX_CONTROL_RATIO = 60
PICI_HORIZONS = (7, 14, 21, 28, 35, 42, 49, 56, 63, 70, 77, 84, 98)
SLOW_SHIP_REPLENISHMENT_START_DAY = 61
FLAT_STAGNANT_SHIPPING_CUTOFF_DAY = 75
SUPPLY_CHANNELS = (
    {
        "channel": "urgent_air",
        "label": "加急空运",
        "arrival_day": 10,
        "window_start": 10,
        "window_end": 19,
    },
    {
        "channel": "standard_air",
        "label": "普通空运",
        "arrival_day": 20,
        "window_start": 20,
        "window_end": 45,
    },
    {
        "channel": "fast_ship",
        "label": "快船",
        "arrival_day": 45,
        "window_start": 46,
        "window_end": 60,
    },
    {
        "channel": "slow_ship",
        "label": "慢船",
        "arrival_day": 60,
        "window_start": 61,
        "window_end": float("inf"),
    },
)

DETAIL_HEADERS = [
    "建议口径",
    "销售部门",
    "销售员",
    "店铺",
    "国家",
    "发货国家",
    "SKU",
    "MSKU",
    "FNSKU",
    "ASIN",
    "品名",
    "销售属性",
    "产品属性",
    "季节属性",
    "MSKU状态",
    "风险类型",
    "风险等级",
    "提示",
    "最早断货天",
    "原始断货天数",
    "模拟后断货天数",
    "关键缺口",
    "控销段",
    "控销天数",
    "控销需求缺口",
    "最高控销比例",
    "控销后剩余缺口",
    "无法仅靠控销覆盖的断货段",
    "加急空运建议件数",
    "普通空运建议件数",
    "快船建议件数",
    "慢船建议件数",
    "PMC建议",
    "采购建议",
    "销售建议",
    "FBA可售",
    "总库存",
    "在途合计",
    "30天需求",
    "日均需求",
    "FBA覆盖天数",
    "提前期",
]

SUMMARY_HEADERS = [
    "建议口径",
    "销售部门",
    "销售员",
    "SKU建议行数",
    "SKU数",
    "断货SKU数",
    "冗余SKU数",
    "控销SKU数",
    "控销总天数",
    "控销需求缺口",
    "控销后剩余缺口",
    "加急空运建议件数",
    "普通空运建议件数",
    "快船建议件数",
    "慢船建议件数",
    "需补货SKU数",
    "最高风险SKU",
    "PMC建议",
    "采购建议",
    "销售建议",
]


def build_recommendation_export_workbook(
    filters: dict[str, Any] | None = None,
    material_code: str | None = None,
    sales_date: str | None = None,
    sales_start_date: str | None = None,
    sales_end_date: str | None = None,
    max_rows: int = 20000,
    connector: StiDatabaseConnector | None = None,
) -> tuple[bytes, dict[str, Any]]:
    rows, meta = build_recommendation_rows(
        filters=filters,
        material_code=material_code,
        sales_date=sales_date,
        sales_start_date=sales_start_date,
        sales_end_date=sales_end_date,
        max_rows=max_rows,
        connector=connector,
    )
    workbook = Workbook()
    overall_sheet = workbook.active
    overall_sheet.title = "整体建议"
    _write_table(overall_sheet, SUMMARY_HEADERS, _aggregate_rows(rows, group_fields=()))
    _write_table(workbook.create_sheet("部门建议"), SUMMARY_HEADERS, _aggregate_rows(rows, group_fields=("sales_department",)))
    _write_table(workbook.create_sheet("个人建议"), SUMMARY_HEADERS, _aggregate_rows(rows, group_fields=("sales_department", "salesman")))
    _write_table(workbook.create_sheet("SKU明细"), DETAIL_HEADERS, [_detail_excel_row(row) for row in rows])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), {**meta, "export_rows": len(rows), "sheet_names": workbook.sheetnames}


def build_recommendation_rows(
    filters: dict[str, Any] | None = None,
    material_code: str | None = None,
    sales_date: str | None = None,
    sales_start_date: str | None = None,
    sales_end_date: str | None = None,
    max_rows: int = 20000,
    connector: StiDatabaseConnector | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    bounded_max_rows = max(1, min(int(max_rows or 20000), 20000))
    summary = get_control_tower_summary(
        material_code=material_code,
        filters=filters or {"order_by": "risk_then_demand"},
        sales_date=sales_date,
        sales_start_date=sales_start_date,
        sales_end_date=sales_end_date,
        page=1,
        page_size=bounded_max_rows,
        max_rows=bounded_max_rows,
        connector=connector,
    )
    rows: list[dict[str, Any]] = []
    for item in summary.items:
        if _is_flat_or_stagnant(item.sales_property):
            recommendation = _build_strategy_recommendation(
                item,
                strategy_key="flat_slow_60",
                strategy_label="当前平滞控销口径",
                target_limit_day=60,
                replenishment_mode="slow_ship_only",
                control_mode="recovery_segmented",
            )
        else:
            recommendation = _build_strategy_recommendation(
                item,
                strategy_key="standard_45",
                strategy_label="当前45天控销口径",
                target_limit_day=45,
            )
        if _is_actionable_row(recommendation):
            rows.append(recommendation)
    return rows, {
        "source_rows": summary.pagination.get("total_count", len(summary.items)),
        "loaded_rows": len(summary.items),
        "sales_period": summary.sales_stat_date,
        "max_rows": bounded_max_rows,
    }


def recommendation_export_filename() -> str:
    return f"库存控销补货建议_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


def _build_strategy_recommendation(
    item: ControlTowerItem,
    strategy_key: str,
    strategy_label: str,
    target_limit_day: int,
    replenishment_mode: str = "standard",
    control_mode: str = "auto",
) -> dict[str, Any]:
    base_summary = _pici_shortage_window_summary(item)
    plan = _build_sales_control_plan(
        item,
        target_limit_day=target_limit_day,
        strategy_key=strategy_key,
        strategy_label=strategy_label,
        replenishment_mode=replenishment_mode,
        control_mode=control_mode,
    )
    control_simulation = _simulation_for_segments(plan["segments"])
    replenishment_windows = _sequential_replenishment_windows(item, control_simulation, replenishment_mode)
    final_simulation = _simulation_with_replenishment(control_simulation, replenishment_windows)
    final_summary = _pici_shortage_window_summary(item, final_simulation)
    control_days = _control_day_count(plan["segments"])
    urgent_air_qty = replenishment_windows["urgent_air"]["suggested_quantity"]
    standard_air_qty = replenishment_windows["standard_air"]["suggested_quantity"]
    fast_qty = replenishment_windows["fast_ship"]["suggested_quantity"]
    slow_qty = replenishment_windows["slow_ship"]["suggested_quantity"]
    return {
        "strategy_key": strategy_key,
        "strategy_label": strategy_label,
        "sales_department": item.sales_department or "未填部门",
        "salesman": item.salesman or "未填销售员",
        "store_name": item.store_name,
        "country_code": item.country_code,
        "shipments_country": item.shipments_country,
        "material_code": item.material_code,
        "msku": item.msku,
        "fnsku": item.fnsku,
        "asin": item.asin,
        "sku_name": item.sku_name,
        "sales_property": item.sales_property,
        "product_property": item.product_property,
        "seasonality": item.seasonality,
        "msku_status": item.msku_status,
        "risk_type": item.risk_type,
        "risk_level": item.risk_level,
        "warning_type": item.warning_type,
        "risk_score": item.risk_score,
        "first_shortage_day": base_summary["first_start_day"],
        "original_shortage_days": base_summary["total_days"],
        "simulated_shortage_days": final_summary["total_days"],
        "pici_key_gap": item.pici_key_gap,
        "control_segments": _format_control_segments(plan["segments"]),
        "control_days": control_days,
        "control_saved_quantity": round(_number(plan["control_saved_quantity"]), 1),
        "max_control_ratio": _max_control_ratio(plan["segments"]),
        "residual_shortage_quantity": round(_number(plan["residual_shortage_quantity"]), 1),
        "unresolved_segments": _format_unresolved_segments(plan["unresolved_segments"]),
        "urgent_air_quantity": urgent_air_qty,
        "standard_air_quantity": standard_air_qty,
        "fast_quantity": fast_qty,
        "slow_quantity": slow_qty,
        "replenishment_mode": replenishment_mode,
        "target_limit_day": target_limit_day,
        "replenishment_text": _format_replenishment_text(
            urgent_air_qty,
            standard_air_qty,
            fast_qty,
            slow_qty,
            replenishment_windows["slow_ship"],
        ),
        "pmc_recommendation": _pmc_recommendation_text(
            urgent_air_qty,
            standard_air_qty,
            fast_qty,
            slow_qty,
            replenishment_windows["slow_ship"],
        ),
        "procurement_recommendation": "",
        "sales_recommendation": _sales_recommendation_text(plan),
        "fba_sellable": item.fba_sellable,
        "total_inventory": item.total_inventory,
        "inbound_total": item.inbound_total,
        "demand_30d": item.demand_30d,
        "daily_demand": item.daily_demand,
        "sellable_days": item.sellable_days,
        "lead_time_days": item.lead_time_days,
        "stockout_active": _is_active_level(item.stockout_risk_level),
        "overstock_active": _is_active_level(item.overstock_risk_level),
        "has_replenishment": bool(urgent_air_qty or standard_air_qty or fast_qty or slow_qty),
        "has_control": bool(control_days or plan["control_saved_quantity"]),
    }


def _sequential_replenishment_windows(
    item: ControlTowerItem,
    control_simulation: dict[str, Any],
    replenishment_mode: str,
) -> dict[str, dict[str, Any]]:
    empty_window = _empty_window()
    if replenishment_mode == "slow_ship_only":
        controlled_summary = _pici_shortage_window_summary(item, control_simulation)
        return {
            "urgent_air": empty_window,
            "standard_air": empty_window,
            "fast_ship": empty_window,
            "slow_ship": _summarize_supply_window(
                controlled_summary["days"],
                SLOW_SHIP_REPLENISHMENT_START_DAY,
                FLAT_STAGNANT_SHIPPING_CUTOFF_DAY,
            ),
        }

    controlled_summary = _pici_shortage_window_summary(item, control_simulation)
    urgent_air_window = controlled_summary["urgent_air_replenishment_window"]
    with_urgent_air = _simulation_with_replenishment(control_simulation, {"urgent_air": urgent_air_window})
    standard_air_window = _pici_shortage_window_summary(item, with_urgent_air)["standard_air_replenishment_window"]
    with_air = _simulation_with_replenishment(with_urgent_air, {"standard_air": standard_air_window})
    fast_window = _pici_shortage_window_summary(item, with_air)["fast_replenishment_window"]
    with_air_fast = _simulation_with_replenishment(with_air, {"fast_ship": fast_window})
    slow_window = _pici_shortage_window_summary(item, with_air_fast)["slow_replenishment_window"]
    return {
        "urgent_air": urgent_air_window,
        "standard_air": standard_air_window,
        "fast_ship": fast_window,
        "slow_ship": slow_window,
    }


def _simulation_for_segments(segments: list[dict[str, Any]]) -> dict[str, Any]:
    return {
        channel["channel"]: {"replenishQuantity": 0}
        for channel in SUPPLY_CHANNELS
    } | {"salesControls": [dict(segment) for segment in segments]}


def _simulation_with_replenishment(
    simulation: dict[str, Any],
    windows_by_channel: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    next_simulation = {
        channel["channel"]: dict((simulation or {}).get(channel["channel"], {"replenishQuantity": 0}))
        for channel in SUPPLY_CHANNELS
    }
    next_simulation["salesControls"] = list((simulation or {}).get("salesControls") or [])
    for channel in SUPPLY_CHANNELS:
        key = channel["channel"]
        if key not in windows_by_channel:
            continue
        quantity = windows_by_channel[key].get("suggested_quantity", 0)
        if quantity:
            next_simulation[key]["replenishQuantity"] = quantity
    return next_simulation


def _build_sales_control_plan(
    item: ControlTowerItem,
    target_limit_day: int,
    strategy_key: str,
    strategy_label: str,
    replenishment_mode: str,
    control_mode: str = "auto",
) -> dict[str, Any]:
    start_day = 1
    base_summary = _pici_shortage_window_summary(item)
    base_window = _summarize_supply_window(base_summary["days"], start_day, target_limit_day)
    plan_meta = {
        "id": strategy_key,
        "title": strategy_label,
        "replenishment_mode": replenishment_mode,
        "target_limit_day": target_limit_day,
    }
    if not base_window["suggested_quantity"]:
        return {
            **plan_meta,
            "segments": [],
            "unresolved_segments": [],
            "control_saved_quantity": 0,
            "residual_shortage_quantity": 0,
            "shortage_quantity": 0,
            "strategy": "none",
        }
    balanced_plan = _solve_balanced_sales_control_plan(item, base_summary["days"], start_day, target_limit_day, MAX_CONTROL_RATIO, base_window)
    segmented_plan = _solve_segmented_sales_control_plan(item, start_day, target_limit_day, MAX_CONTROL_RATIO, base_window)
    if control_mode == "recovery_segmented":
        return {
            **_solve_recovery_segmented_sales_control_plan(
                item,
                start_day,
                target_limit_day,
                MAX_CONTROL_RATIO,
                base_window,
            ),
            **plan_meta,
        }
    selected = balanced_plan if balanced_plan and _should_use_balanced_control_plan(balanced_plan, segmented_plan) else segmented_plan
    return {**selected, **plan_meta}


def _solve_balanced_sales_control_plan(
    item: ControlTowerItem,
    base_days: list[dict[str, Any]],
    start_day: int,
    target_limit_day: int,
    max_control_ratio: int,
    base_window: dict[str, Any],
) -> dict[str, Any] | None:
    base_segments = _shortage_segments_in_window(base_days, start_day, target_limit_day)
    if not base_segments:
        return None
    end_day = base_segments[-1]["end_day"]

    def build_simulation(control_ratio: float) -> dict[str, Any]:
        return _simulation_for_segments([{"start_day": start_day, "end_day": end_day, "control_ratio": control_ratio}])

    max_summary = _pici_shortage_window_summary(item, build_simulation(max_control_ratio))
    max_window = _summarize_supply_window(max_summary["days"], start_day, end_day)
    if max_window["shortage_quantity"] > 0:
        shortage_segments = _shortage_segments_in_window(max_summary["days"], start_day, end_day)
        return {
            "segments": [
                {
                    "start_day": start_day,
                    "end_day": end_day,
                    "target_start_day": start_day,
                    "target_end_day": end_day,
                    "control_ratio": max_control_ratio,
                    "control_saved_quantity": max_window["control_saved_quantity"],
                    "residual_shortage_quantity": max_window["shortage_quantity"],
                    "residual_shortage_segments": shortage_segments,
                    "unresolved": True,
                }
            ],
            "unresolved_segments": [
                {
                    "start_day": start_day,
                    "end_day": end_day,
                    "shortage_quantity": max_window["shortage_quantity"],
                    "shortage_segments": shortage_segments,
                    "reason": "最高60%平滑控销仍无法覆盖。",
                }
            ],
            "control_saved_quantity": max_window["control_saved_quantity"],
            "residual_shortage_quantity": max_window["shortage_quantity"],
            "shortage_quantity": base_window["shortage_quantity"],
            "target_limit_day": target_limit_day,
            "strategy": "balanced",
        }

    lower = 0.0
    upper = float(max_control_ratio)
    for _ in range(10):
        middle = (lower + upper) / 2
        summary = _pici_shortage_window_summary(item, build_simulation(middle))
        window = _summarize_supply_window(summary["days"], start_day, end_day)
        if window["shortage_quantity"] > 0:
            lower = middle
        else:
            upper = middle
    control_ratio = _ceil(upper)
    final_summary = _pici_shortage_window_summary(item, build_simulation(control_ratio))
    final_window = _summarize_supply_window(final_summary["days"], start_day, end_day)
    return {
        "segments": [
            {
                "start_day": start_day,
                "end_day": end_day,
                "target_start_day": start_day,
                "target_end_day": end_day,
                "control_ratio": control_ratio,
                "control_saved_quantity": final_window["control_saved_quantity"],
                "residual_shortage_quantity": final_window["shortage_quantity"],
                "residual_shortage_segments": [],
                "unresolved": False,
            }
        ],
        "unresolved_segments": [],
        "control_saved_quantity": final_window["control_saved_quantity"],
        "residual_shortage_quantity": final_window["shortage_quantity"],
        "shortage_quantity": base_window["shortage_quantity"],
        "target_limit_day": target_limit_day,
        "strategy": "balanced",
    }


def _solve_segmented_sales_control_plan(
    item: ControlTowerItem,
    start_day: int,
    target_limit_day: int,
    max_control_ratio: int,
    base_window: dict[str, Any],
) -> dict[str, Any]:
    segments: list[dict[str, Any]] = []
    unresolved_segments: list[dict[str, Any]] = []
    handled_until = start_day - 1
    for _ in range(20):
        current_summary = _pici_shortage_window_summary(item, _simulation_for_segments(segments))
        next_segment = next(
            (
                segment
                for segment in _shortage_segments_in_window(current_summary["days"], start_day, target_limit_day)
                if segment["end_day"] > handled_until
            ),
            None,
        )
        if not next_segment:
            break
        control_start_day = max(start_day, handled_until + 1)
        solved = _solve_sales_control_segment(
            item,
            segments,
            control_start_day,
            next_segment["start_day"],
            next_segment["end_day"],
            max_control_ratio,
        )
        if solved:
            segments.append(solved)
        if solved and solved.get("unresolved"):
            unresolved_segments.append(
                {
                    "start_day": solved["target_start_day"],
                    "end_day": solved["target_end_day"],
                    "shortage_quantity": solved["residual_shortage_quantity"],
                    "shortage_segments": solved["residual_shortage_segments"],
                    "reason": "最高60%控销仍无法覆盖该断货段",
                }
            )
        handled_until = next_segment["end_day"]

    final_summary = _pici_shortage_window_summary(item, _simulation_for_segments(segments))
    final_window = _summarize_supply_window(final_summary["days"], start_day, target_limit_day)
    return {
        "segments": segments,
        "unresolved_segments": unresolved_segments,
        "control_saved_quantity": _round_to_one(sum(_number(segment.get("control_saved_quantity")) for segment in segments)),
        "residual_shortage_quantity": final_window["shortage_quantity"],
        "shortage_quantity": base_window["shortage_quantity"],
        "target_limit_day": target_limit_day,
        "strategy": "segmented",
    }


def _solve_recovery_segmented_sales_control_plan(
    item: ControlTowerItem,
    start_day: int,
    target_limit_day: int,
    max_control_ratio: int,
    base_window: dict[str, Any],
) -> dict[str, Any]:
    segments: list[dict[str, Any]] = []
    unresolved_segments: list[dict[str, Any]] = []
    handled_until = start_day - 1
    for _ in range(20):
        current_summary = _pici_shortage_window_summary(item, _simulation_for_segments(segments))
        next_segment = next(
            (
                segment
                for segment in _shortage_segments_in_window(current_summary["days"], start_day, target_limit_day)
                if segment["end_day"] > handled_until
            ),
            None,
        )
        if not next_segment:
            break
        recovery_day = _first_recovered_supply_day(current_summary["days"], handled_until, next_segment["start_day"])
        control_start_day = recovery_day or max(start_day, handled_until + 1)
        solved = _solve_sales_control_segment(
            item,
            segments,
            control_start_day,
            next_segment["start_day"],
            next_segment["end_day"],
            max_control_ratio,
        )
        if solved:
            segments.append(solved)
        if solved and solved.get("unresolved"):
            unresolved_segments.append(
                {
                    "start_day": solved["target_start_day"],
                    "end_day": solved["target_end_day"],
                    "shortage_quantity": solved["residual_shortage_quantity"],
                    "shortage_segments": solved["residual_shortage_segments"],
                    "reason": "最高60%分段控销仍无法覆盖该断货段",
                }
            )
        handled_until = next_segment["end_day"]

    final_summary = _pici_shortage_window_summary(item, _simulation_for_segments(segments))
    final_window = _summarize_supply_window(final_summary["days"], start_day, target_limit_day)
    return {
        "segments": segments,
        "unresolved_segments": unresolved_segments,
        "control_saved_quantity": _round_to_one(sum(_number(segment.get("control_saved_quantity")) for segment in segments)),
        "residual_shortage_quantity": final_window["shortage_quantity"],
        "shortage_quantity": base_window["shortage_quantity"],
        "target_limit_day": target_limit_day,
        "strategy": "recovery_segmented",
    }


def _first_recovered_supply_day(days: list[dict[str, Any]], handled_until: int, target_start_day: int) -> int | None:
    return next(
        (
            int(day["day"])
            for day in days or []
            if int(day.get("day") or 0) > handled_until
            and int(day.get("day") or 0) <= target_start_day
            and day.get("status") != "shortage"
        ),
        None,
    )


def _solve_sales_control_segment(
    item: ControlTowerItem,
    existing_segments: list[dict[str, Any]],
    control_start_day: int,
    target_start_day: int,
    target_end_day: int,
    max_control_ratio: int,
) -> dict[str, Any]:
    def build_simulation(control_ratio: float) -> dict[str, Any]:
        return _simulation_for_segments(
            [
                *existing_segments,
                {
                    "start_day": control_start_day,
                    "end_day": target_end_day,
                    "control_ratio": control_ratio,
                },
            ]
        )

    max_summary = _pici_shortage_window_summary(item, build_simulation(max_control_ratio))
    max_target_window = _summarize_supply_window(max_summary["days"], target_start_day, target_end_day)
    if max_target_window["shortage_quantity"] > 0:
        max_control_window = _summarize_supply_window(max_summary["days"], control_start_day, target_end_day)
        return {
            "start_day": control_start_day,
            "end_day": target_end_day,
            "target_start_day": target_start_day,
            "target_end_day": target_end_day,
            "control_ratio": max_control_ratio,
            "control_saved_quantity": max_control_window["control_saved_quantity"],
            "residual_shortage_quantity": max_target_window["shortage_quantity"],
            "residual_shortage_segments": _shortage_segments_in_window(max_summary["days"], target_start_day, target_end_day),
            "unresolved": True,
        }

    lower = 0.0
    upper = float(max_control_ratio)
    for _ in range(10):
        middle = (lower + upper) / 2
        summary = _pici_shortage_window_summary(item, build_simulation(middle))
        target_window = _summarize_supply_window(summary["days"], target_start_day, target_end_day)
        if target_window["shortage_quantity"] > 0:
            lower = middle
        else:
            upper = middle
    control_ratio = _ceil(upper)
    final_summary = _pici_shortage_window_summary(item, build_simulation(control_ratio))
    control_window = _summarize_supply_window(final_summary["days"], control_start_day, target_end_day)
    target_window = _summarize_supply_window(final_summary["days"], target_start_day, target_end_day)
    return {
        "start_day": control_start_day,
        "end_day": target_end_day,
        "target_start_day": target_start_day,
        "target_end_day": target_end_day,
        "control_ratio": control_ratio,
        "control_saved_quantity": control_window["control_saved_quantity"],
        "residual_shortage_quantity": target_window["shortage_quantity"],
        "residual_shortage_segments": [],
        "unresolved": False,
    }


def _pici_shortage_window_summary(item: ControlTowerItem, simulation: dict[str, Any] | None = None) -> dict[str, Any]:
    entries = []
    for key, value in (item.pici_gap_values or {}).items():
        try:
            horizon = int(str(key).split("_")[-1])
        except ValueError:
            continue
        parsed = _parse_pici_value(value)
        if horizon and parsed:
            entries.append({"horizon": horizon, **parsed})
    entries.sort(key=lambda entry: entry["horizon"])
    previous_horizon = 0
    previous_available = 0.0
    previous_forecast = 0.0
    base_pool = 0.0
    simulation_pools: list[dict[str, Any]] = []
    total_days = 0
    segments: list[str] = []
    days: list[dict[str, Any]] = []
    active_shortage_start: int | None = None
    first_start_day: int | None = None
    config = _normalize_supply_simulation(simulation)

    for entry in entries:
        interval_days = max(entry["horizon"] - previous_horizon, 0)
        interval_forecast = max(entry["forecast"] - previous_forecast, 0)
        interval_supply = max(entry["available"] - previous_available, 0)
        interval_start_day = previous_horizon + 1
        interval_end_day = entry["horizon"]
        if interval_days > 0:
            daily_forecast = interval_forecast / interval_days
            base_pool += interval_supply
            for day in range(interval_start_day, interval_end_day + 1):
                arrival = _simulated_arrival_for_day(config, day)
                if arrival["quantity"] > 0:
                    simulation_pools.append({"channel": arrival["channel"], "quantity": arrival["quantity"]})

                control_ratio = _control_ratio_for_day(config, day)
                effective_forecast = daily_forecast * (1 - control_ratio / 100)
                control_saved_quantity = max(daily_forecast - effective_forecast, 0)
                remaining_demand = effective_forecast
                replenished_quantity = 0.0

                base_used = min(base_pool, remaining_demand)
                base_pool -= base_used
                remaining_demand -= base_used

                while remaining_demand > 0 and simulation_pools:
                    pool = simulation_pools[0]
                    used_quantity = min(pool["quantity"], remaining_demand)
                    replenished_quantity += used_quantity
                    pool["quantity"] -= used_quantity
                    remaining_demand -= used_quantity
                    if pool["quantity"] <= 0.000001:
                        simulation_pools.pop(0)

                shortage_quantity = max(remaining_demand, 0)
                is_covered = daily_forecast <= 0 or remaining_demand <= 0.000001
                status = "replenished" if is_covered and replenished_quantity > 0 else "ok" if is_covered else "shortage"
                days.append(
                    {
                        "day": day,
                        "status": status,
                        "control_ratio": control_ratio,
                        "forecast": effective_forecast,
                        "original_forecast": daily_forecast,
                        "control_saved_quantity": control_saved_quantity,
                        "shortage_quantity": shortage_quantity,
                        "replenished_quantity": replenished_quantity,
                    }
                )

                if is_covered:
                    if active_shortage_start is not None:
                        segments.append(_format_shortage_segment(active_shortage_start, day - active_shortage_start))
                        active_shortage_start = None
                else:
                    total_days += 1
                    first_start_day = day if first_start_day is None else min(first_start_day, day)
                    active_shortage_start = active_shortage_start or day
        previous_horizon = entry["horizon"]
        previous_available = entry["available"]
        previous_forecast = entry["forecast"]

    if active_shortage_start is not None:
        latest_day = days[-1]["day"] if days else active_shortage_start
        segments.append(_format_shortage_segment(active_shortage_start, latest_day - active_shortage_start + 1))

    return {
        "total_days": total_days,
        "first_start_day": first_start_day or item.pici_first_shortage_days or 0,
        "segments": segments,
        "days": days,
        "urgent_air_replenishment_window": _summarize_supply_window(days, 10, 19),
        "standard_air_replenishment_window": _summarize_supply_window(days, 20, 45),
        "fast_replenishment_window": _summarize_supply_window(days, 46, 60),
        "slow_replenishment_window": _summarize_supply_window(days, SLOW_SHIP_REPLENISHMENT_START_DAY, float("inf")),
    }


def _normalize_supply_simulation(simulation: dict[str, Any] | None) -> dict[str, Any]:
    source = simulation or {}
    result: dict[str, Any] = {}
    for channel in SUPPLY_CHANNELS:
        values = source.get(channel["channel"], {}) or {}
        if channel["channel"] == "standard_air" and not values:
            values = source.get("air_or_urgent_transfer", {}) or {}
        result[channel["channel"]] = {
            **channel,
            "replenish_quantity": max(_number(values.get("replenishQuantity") or values.get("replenish_quantity")), 0),
        }
    result["sales_controls"] = _normalize_sales_control_segments(source.get("salesControls") or source.get("sales_controls") or [])
    return result


def _normalize_sales_control_segments(raw_segments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    segments = []
    for control in raw_segments or []:
        start_day = _number(control.get("startDay") or control.get("start_day"))
        end_day = _number(control.get("endDay") or control.get("end_day"))
        control_ratio = max(0, min(_number(control.get("controlRatio") or control.get("control_ratio")), MAX_CONTROL_RATIO))
        if start_day <= 0 or end_day <= 0 or not control_ratio:
            continue
        segments.append(
            {
                "start_day": int(min(start_day, end_day)),
                "end_day": int(max(start_day, end_day)),
                "control_ratio": control_ratio,
            }
        )
    return sorted(segments, key=lambda segment: (segment["start_day"], segment["end_day"]))


def _simulated_arrival_for_day(config: dict[str, Any], day: int) -> dict[str, Any]:
    for channel in SUPPLY_CHANNELS:
        values = config[channel["channel"]]
        if values["arrival_day"] == day and values["replenish_quantity"] > 0:
            return {"channel": values["channel"], "quantity": values["replenish_quantity"]}
    return {"channel": "", "quantity": 0}


def _control_ratio_for_day(config: dict[str, Any], day: int) -> float:
    ratios = [
        control["control_ratio"]
        for control in config.get("sales_controls", [])
        if day >= control["start_day"] and day <= control["end_day"]
    ]
    return min(MAX_CONTROL_RATIO, max([0, *ratios]))


def _summarize_supply_window(days: list[dict[str, Any]], start_day: int, end_day: float) -> dict[str, Any]:
    window_days = [day for day in days if day["day"] >= start_day and day["day"] <= end_day]
    shortage_quantity = sum(max(_number(day.get("shortage_quantity")), 0) for day in window_days)
    control_saved_quantity = sum(max(_number(day.get("control_saved_quantity")), 0) for day in window_days)
    demand_quantity = sum(max(_number(day.get("forecast")), 0) for day in window_days)
    original_demand_quantity = sum(max(_number(day.get("original_forecast")), 0) for day in window_days)
    return {
        "start_day": start_day,
        "end_day": end_day,
        "shortage_days": sum(1 for day in window_days if day.get("status") == "shortage"),
        "shortage_quantity": _round_to_one(shortage_quantity),
        "suggested_quantity": _ceil(shortage_quantity),
        "control_saved_quantity": _round_to_one(control_saved_quantity),
        "demand_quantity": _round_to_one(demand_quantity),
        "original_demand_quantity": _round_to_one(original_demand_quantity),
    }


def _shortage_segments_in_window(days: list[dict[str, Any]], start_day: int, end_day: int) -> list[dict[str, Any]]:
    segments = []
    active_start: int | None = None
    shortage_quantity = 0.0
    for day in days or []:
        day_number = int(day.get("day") or 0)
        if day_number < start_day or day_number > end_day:
            continue
        if day.get("status") == "shortage":
            active_start = active_start or day_number
            shortage_quantity += max(_number(day.get("shortage_quantity")), 0)
            continue
        if active_start is not None:
            segments.append({"start_day": active_start, "end_day": day_number - 1, "shortage_quantity": _round_to_one(shortage_quantity)})
            active_start = None
            shortage_quantity = 0.0
    if active_start is not None:
        latest_day = min(end_day, int((days or [{}])[-1].get("day") or end_day))
        segments.append({"start_day": active_start, "end_day": latest_day, "shortage_quantity": _round_to_one(shortage_quantity)})
    return segments


def _should_use_balanced_control_plan(balanced_plan: dict[str, Any], segmented_plan: dict[str, Any]) -> bool:
    balanced_max = _max_control_ratio(balanced_plan.get("segments", []))
    segmented_max = _max_control_ratio(segmented_plan.get("segments", []))
    if not balanced_plan.get("residual_shortage_quantity") and segmented_plan.get("residual_shortage_quantity"):
        return True
    if balanced_plan.get("residual_shortage_quantity") and not segmented_plan.get("residual_shortage_quantity"):
        return False
    if len(segmented_plan.get("segments", [])) > 1 and balanced_max <= segmented_max:
        return True
    return balanced_max + 6 < segmented_max


def _aggregate_rows(rows: list[dict[str, Any]], group_fields: tuple[str, ...]) -> list[list[Any]]:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = tuple(row.get(field, "") for field in group_fields) + (row.get("strategy_label", ""),)
        grouped[key].append(row)

    output = []
    for key, group_rows in sorted(grouped.items(), key=lambda item: tuple(str(value) for value in item[0])):
        strategy_label = key[-1]
        department = key[0] if "sales_department" in group_fields else "全部"
        salesman = key[1] if "salesman" in group_fields else "全部"
        sku_codes = {row["material_code"] for row in group_rows}
        stockout_skus = {row["material_code"] for row in group_rows if row["stockout_active"]}
        overstock_skus = {row["material_code"] for row in group_rows if row["overstock_active"]}
        control_skus = {row["material_code"] for row in group_rows if row["has_control"]}
        replenishment_skus = {row["material_code"] for row in group_rows if row["has_replenishment"]}
        highest = max(group_rows, key=lambda row: row.get("risk_score", 0), default={})
        urgent_air_qty = sum(_number(row["urgent_air_quantity"]) for row in group_rows)
        standard_air_qty = sum(_number(row["standard_air_quantity"]) for row in group_rows)
        fast_qty = sum(_number(row["fast_quantity"]) for row in group_rows)
        slow_qty = sum(_number(row["slow_quantity"]) for row in group_rows)
        control_saved = sum(_number(row["control_saved_quantity"]) for row in group_rows)
        residual_shortage = sum(_number(row["residual_shortage_quantity"]) for row in group_rows)
        control_days = sum(_number(row["control_days"]) for row in group_rows)
        output.append(
            [
                strategy_label,
                department,
                salesman,
                len(group_rows),
                len(sku_codes),
                len(stockout_skus),
                len(overstock_skus),
                len(control_skus),
                int(control_days),
                _round_to_one(control_saved),
                _round_to_one(residual_shortage),
                int(urgent_air_qty),
                int(standard_air_qty),
                int(fast_qty),
                int(slow_qty),
                len(replenishment_skus),
                highest.get("material_code", ""),
                _aggregate_pmc_recommendation_text(urgent_air_qty, standard_air_qty, fast_qty, slow_qty, residual_shortage),
                "",
                _aggregate_sales_recommendation_text(control_skus, control_saved),
            ]
        )
    return output


def _aggregate_pmc_recommendation_text(
    urgent_air_qty: float,
    standard_air_qty: float,
    fast_qty: float,
    slow_qty: float,
    residual_shortage: float,
) -> str:
    parts = []
    if urgent_air_qty or standard_air_qty or fast_qty or slow_qty:
        parts.append(f"加急空运 {int(urgent_air_qty)} 件，普通空运 {int(standard_air_qty)} 件，快船 {int(fast_qty)} 件，慢船 {int(slow_qty)} 件")
    if residual_shortage:
        parts.append(f"控销后仍缺 {_format_number(residual_shortage)} 件，需补货覆盖")
    return "；".join(parts)


def _aggregate_sales_recommendation_text(control_skus: set[str], control_saved: float) -> str:
    if not control_skus:
        return ""
    return f"控销 {len(control_skus)} 个SKU，需求缺口 {_format_number(control_saved)} 件"


def _detail_excel_row(row: dict[str, Any]) -> list[Any]:
    return [
        row["strategy_label"],
        row["sales_department"],
        row["salesman"],
        row["store_name"],
        row["country_code"],
        row["shipments_country"],
        row["material_code"],
        row["msku"],
        row["fnsku"],
        row["asin"],
        row["sku_name"],
        row["sales_property"],
        row["product_property"],
        row["seasonality"],
        row["msku_status"],
        row["risk_type"],
        row["risk_level"],
        row["warning_type"],
        row["first_shortage_day"] or "",
        row["original_shortage_days"],
        row["simulated_shortage_days"],
        row["pici_key_gap"],
        row["control_segments"],
        row["control_days"],
        row["control_saved_quantity"],
        row["max_control_ratio"],
        row["residual_shortage_quantity"],
        row["unresolved_segments"],
        row["urgent_air_quantity"],
        row["standard_air_quantity"],
        row["fast_quantity"],
        row["slow_quantity"],
        row["pmc_recommendation"],
        row["procurement_recommendation"],
        row["sales_recommendation"],
        row["fba_sellable"],
        row["total_inventory"],
        row["inbound_total"],
        row["demand_30d"],
        row["daily_demand"],
        row["sellable_days"] if row["sellable_days"] is not None else "",
        row["lead_time_days"],
    ]


def _write_table(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, sheet.max_row)}"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, len(headers) + 1):
        letter = get_column_letter(column)
        width = 14
        header = headers[column - 1]
        if header in {"品名", "PMC建议", "采购建议", "销售建议", "无法仅靠控销覆盖的断货段"}:
            width = 34
        elif header in {"控销段", "提示", "关键缺口"}:
            width = 24
        elif header in {"SKU", "MSKU", "FNSKU", "ASIN"}:
            width = 18
        sheet.column_dimensions[letter].width = width


def _is_actionable_row(row: dict[str, Any]) -> bool:
    return bool(
        row.get("stockout_active")
        or row.get("overstock_active")
        or row.get("risk_type") == "anomaly"
        or row.get("has_control")
        or row.get("has_replenishment")
    )


def _has_supply_action(row: dict[str, Any]) -> bool:
    return bool(
        row.get("stockout_active")
        or row.get("original_shortage_days")
        or row.get("has_control")
        or row.get("has_replenishment")
    )


def _pmc_recommendation_text(
    urgent_air_qty: int,
    standard_air_qty: int,
    fast_qty: int,
    slow_qty: int,
    slow_window: dict[str, Any] | None = None,
) -> str:
    return _format_replenishment_text(urgent_air_qty, standard_air_qty, fast_qty, slow_qty, slow_window)


def _sales_recommendation_text(plan: dict[str, Any]) -> str:
    if plan.get("segments"):
        return _format_control_segments(plan["segments"])
    return ""


def _format_control_segments(segments: list[dict[str, Any]]) -> str:
    if not segments:
        return ""
    return "；".join(
        f"第{_format_number(segment['start_day'])}-{_format_number(segment['end_day'])}天 控{_format_number(segment['control_ratio'])}%"
        for segment in segments
    )


def _format_unresolved_segments(segments: list[dict[str, Any]]) -> str:
    if not segments:
        return ""
    parts = []
    for segment in segments:
        shortage_segments = segment.get("shortage_segments") or []
        shortage_text = "、".join(
            f"第{_format_number(item['start_day'])}-{_format_number(item['end_day'])}天缺{_format_number(item['shortage_quantity'])}件"
            for item in shortage_segments
        )
        parts.append(shortage_text or f"第{_format_number(segment['start_day'])}-{_format_number(segment['end_day'])}天缺{_format_number(segment['shortage_quantity'])}件")
    return "；".join(parts)


def _format_replenishment_text(
    urgent_air_qty: float,
    standard_air_qty: float,
    fast_qty: float,
    slow_qty: float,
    slow_window: dict[str, Any] | None = None,
) -> str:
    parts = []
    if urgent_air_qty:
        parts.append(f"加急空运{int(urgent_air_qty)}件(第10天到，覆盖10-19天)")
    if standard_air_qty:
        parts.append(f"普通空运{int(standard_air_qty)}件(第20天到，覆盖20-45天)")
    if fast_qty:
        parts.append(f"快船{int(fast_qty)}件(第45天到，覆盖46-60天)")
    if slow_qty:
        parts.append(f"慢船{int(slow_qty)}件(第60天到，覆盖{_format_supply_window_label(slow_window, '61天后')})")
    return "；".join(parts)


def _format_supply_window_label(window: dict[str, Any] | None, fallback: str) -> str:
    if not window:
        return fallback
    start_day = _number(window.get("start_day"))
    end_day = _number(window.get("end_day"))
    if not start_day:
        return fallback
    if end_day == float("inf") or not end_day:
        return f"{_format_number(start_day)}天后"
    return f"{_format_number(start_day)}-{_format_number(end_day)}天"


def _control_day_count(segments: list[dict[str, Any]]) -> int:
    days: set[int] = set()
    for segment in segments:
        start_day = int(segment.get("start_day") or 0)
        end_day = int(segment.get("end_day") or 0)
        if start_day <= 0 or end_day <= 0:
            continue
        days.update(range(min(start_day, end_day), max(start_day, end_day) + 1))
    return len(days)


def _max_control_ratio(segments: list[dict[str, Any]]) -> float:
    return max([0, *[_number(segment.get("control_ratio")) for segment in segments]])


def _parse_pici_value(value: Any) -> dict[str, float] | None:
    match = re.match(r"^\s*([\d,]+(?:\.\d+)?)\s*/\s*([\d,]+(?:\.\d+)?)\s*\((-?[\d,]+(?:\.\d+)?)\)", str(value or ""))
    if not match:
        return None
    return {
        "available": _number(match.group(1).replace(",", "")),
        "forecast": _number(match.group(2).replace(",", "")),
        "gap": _number(match.group(3).replace(",", "")),
    }


def _format_shortage_segment(start_day: int, shortage_days: int) -> str:
    start_label = "今天起" if start_day <= 0 else f"第 {_format_number(start_day)} 天起"
    return f"{start_label}断 {_format_number(shortage_days)} 天"


def _is_flat_or_stagnant(value: str) -> bool:
    text = str(value or "").strip()
    return "平" in text or "滞" in text


def _is_active_level(value: str) -> bool:
    return str(value or "").lower() in {"critical", "high", "medium", "low"}


def _empty_window() -> dict[str, Any]:
    return {
        "start_day": 0,
        "end_day": 0,
        "shortage_days": 0,
        "shortage_quantity": 0,
        "suggested_quantity": 0,
        "control_saved_quantity": 0,
        "demand_quantity": 0,
        "original_demand_quantity": 0,
    }


def _ceil(value: float) -> int:
    numeric = float(value or 0)
    integer = int(numeric)
    return integer if numeric == integer else integer + 1


def _round_to_one(value: float) -> float:
    return round(_number(value), 1)


def _number(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _format_number(value: Any) -> str:
    number = _number(value)
    if number == int(number):
        return str(int(number))
    return f"{number:.1f}".rstrip("0").rstrip(".")
