from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from pmc_agent.connectors.database import StiDatabaseConnector
from pmc_agent.control_tower import ControlTowerItem, get_control_tower_summary
from pmc_agent.sku_diagnosis import prepare_sku_diagnosis_payload


PICI_HORIZONS = (7, 14, 21, 28, 35, 42, 49, 56, 63, 70, 77, 84, 98)
ROOT_CAUSE_LIMIT = 5
CHART_IMAGE_MAX_WIDTH = 980
CHART_ASSET_DIR = Path(__file__).resolve().parent / "assets" / "sku_investigation_charts"
CHART_IMAGE_SPECS = (
    ("断货风险图", CHART_ASSET_DIR / "stockout_risk_timeline.png"),
    ("销量预估详情图", CHART_ASSET_DIR / "sales_forecast_detail.png"),
)
STOCKOUT_REASON_TYPES = {
    "oversell",
    "planning_anomaly",
    "supply_anomaly",
    "logistics_delay",
    "procurement_delay",
    "inventory_position",
    "inventory",
    "stockout",
    "logistics",
    "sales",
    "forecast",
    "forecast_ad",
    "plan",
}
OVERSTOCK_REASON_TYPES = {
    "low_sell",
    "planning_anomaly",
    "supply_anomaly",
    "inventory_position",
    "inventory",
    "overstock",
    "sales",
    "forecast",
    "forecast_ad",
    "plan",
}

RISK_TYPE_LABELS = {
    "stockout": "断货",
    "overstock": "冗余",
    "anomaly": "异常",
    "healthy": "正常",
}
RISK_LEVEL_LABELS = {
    "critical": "紧急",
    "high": "高",
    "medium": "中",
    "low": "低",
    "normal": "正常",
}
ROOT_CAUSE_TYPE_LABELS = {
    "oversell": "超卖",
    "low_sell": "低卖",
    "planning_anomaly": "计划异常",
    "supply_anomaly": "供应异常",
    "logistics_delay": "物流延期",
    "procurement_delay": "采购延期",
    "inventory_position": "库存位置",
    "inventory": "库存明细",
    "stockout": "断货库存",
    "logistics": "物流在途",
    "sales": "销售异常",
    "forecast": "预估异常",
    "forecast_ad": "预测/广告",
    "overstock": "冗余库存",
    "data_quality": "数据异常",
    "monitor": "监控",
    "plan": "计划异常",
}
EXPORT_IGNORED_ROOT_CAUSE_TYPES = {"data_quality"}

DEPARTMENT_SECTION_DEFINITIONS = [
    {
        "key": "sales",
        "title": "销售方向",
        "types": {"oversell", "low_sell", "sales", "forecast", "forecast_ad"},
    },
    {
        "key": "supply",
        "title": "供应方向",
        "types": {"supply_anomaly", "logistics_delay", "procurement_delay", "logistics", "planning_anomaly", "plan"},
    },
    {
        "key": "inventory",
        "title": "库存明细",
        "types": {"inventory_position", "inventory", "stockout", "overstock"},
    },
]

DETAIL_HEADERS = [
    "销售部门",
    "销售员",
    "产品经理",
    "店铺",
    "国家",
    "发货国家",
    "账号",
    "SKU",
    "MSKU",
    "FNSKU",
    "ASIN",
    "品名",
    "销售属性",
    "产品属性",
    "季节属性",
    "MSKU状态",
    "MSKU生命周期",
    "风险类型",
    "整体风险等级",
    "风险分",
    "异常提示",
    "断货分段",
    "销售方向",
    "供应方向",
    "库存明细",
    "销售提示",
]

META_HEADERS = ["项目", "值"]


def build_sku_investigation_export_workbook(
    filters: dict[str, Any] | None = None,
    material_code: str | None = None,
    sales_date: str | None = None,
    sales_start_date: str | None = None,
    sales_end_date: str | None = None,
    max_rows: int = 20000,
    connector: StiDatabaseConnector | None = None,
) -> tuple[bytes, dict[str, Any]]:
    rows, meta = build_sku_investigation_rows(
        filters=filters,
        material_code=material_code,
        sales_date=sales_date,
        sales_start_date=sales_start_date,
        sales_end_date=sales_end_date,
        max_rows=max_rows,
        connector=connector,
    )
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "SKU明细排查"
    _write_table(detail_sheet, DETAIL_HEADERS, [_detail_excel_row(row) for row in rows])
    chart_image_count = _write_chart_screenshot_sheet(workbook)
    meta_with_charts = {**meta, "chart_image_count": chart_image_count}
    _write_table(workbook.create_sheet("导出说明"), META_HEADERS, _meta_rows(meta_with_charts))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue(), {**meta_with_charts, "export_rows": len(rows), "sheet_names": workbook.sheetnames}


def build_sku_investigation_rows(
    filters: dict[str, Any] | None = None,
    material_code: str | None = None,
    sales_date: str | None = None,
    sales_start_date: str | None = None,
    sales_end_date: str | None = None,
    max_rows: int = 20000,
    connector: StiDatabaseConnector | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    bounded_max_rows = max(1, min(int(max_rows or 20000), 20000))
    active_connector = connector or StiDatabaseConnector()
    summary = get_control_tower_summary(
        material_code=material_code,
        filters=filters or {"order_by": "risk_then_demand"},
        sales_date=sales_date,
        sales_start_date=sales_start_date,
        sales_end_date=sales_end_date,
        page=1,
        page_size=bounded_max_rows,
        max_rows=bounded_max_rows,
        connector=active_connector,
    )
    rows = [
        _build_investigation_row(
            item,
            connector=active_connector,
            sales_start_date=summary.sales_start_date,
            sales_end_date=summary.sales_end_date,
        )
        for item in summary.items
    ]
    return rows, {
        "source_rows": summary.pagination.get("total_count", len(summary.items)),
        "loaded_rows": len(summary.items),
        "sales_period": summary.sales_stat_date,
        "sales_start_date": summary.sales_start_date,
        "sales_end_date": summary.sales_end_date,
        "max_rows": bounded_max_rows,
        "filters": filters or {},
        "material_code": material_code or "",
    }


def sku_investigation_export_filename() -> str:
    return f"当前SKU明细排查_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"


def _build_investigation_row(
    item: ControlTowerItem,
    *,
    connector: StiDatabaseConnector | None = None,
    sales_start_date: str | None = None,
    sales_end_date: str | None = None,
) -> dict[str, Any]:
    payload = {
        **asdict(item),
        "sales_start_date": sales_start_date,
        "sales_end_date": sales_end_date,
    }
    enriched_payload, diagnosis_data = prepare_sku_diagnosis_payload(payload, connector=connector)
    stockout_metrics = diagnosis_data.get("stockout", {}).get("metrics", {})
    root_causes = diagnosis_data.get("root_cause_analysis", [])
    department_sections = _department_root_cause_sections(root_causes)
    return {
        "item": enriched_payload,
        "diagnosis": diagnosis_data,
        "risk_summary": _format_risk_summary(enriched_payload, diagnosis_data),
        "attribution": _join_lines(diagnosis_data.get("attribution")),
        "root_cause_summary": _format_root_causes(root_causes),
        "department_sections": department_sections,
        "stockout_reason_summary": _format_stockout_reasons(enriched_payload, diagnosis_data, root_causes),
        "root_causes": root_causes[:ROOT_CAUSE_LIMIT] if isinstance(root_causes, list) else [],
        "shortage_days": stockout_metrics.get("shortage_days"),
    }


def _detail_excel_row(row: dict[str, Any]) -> list[Any]:
    item = row["item"]
    sections = row.get("department_sections") if isinstance(row.get("department_sections"), dict) else {}
    return [
        item.get("sales_department"),
        item.get("salesman"),
        item.get("product_manager"),
        item.get("store_name"),
        item.get("country_code"),
        item.get("shipments_country"),
        item.get("seller_id"),
        item.get("material_code"),
        item.get("msku"),
        item.get("fnsku"),
        item.get("asin"),
        item.get("sku_name"),
        item.get("sales_property"),
        item.get("product_property"),
        item.get("seasonality"),
        item.get("msku_status"),
        item.get("msku_life_process"),
        _label(RISK_TYPE_LABELS, item.get("risk_type")),
        _label(RISK_LEVEL_LABELS, item.get("risk_level")),
        item.get("risk_score"),
        item.get("warning_type"),
        _department_section_text(sections, "stockout_segments"),
        _department_section_text(sections, "sales"),
        _department_section_text(sections, "supply"),
        _department_section_text(sections, "inventory"),
        _format_sales_signal_text(row.get("diagnosis")),
    ]


def _write_table(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, sheet.max_row)}"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, len(headers) + 1):
        header = headers[column - 1]
        width = 14
        if header in {"断货分段", "销售方向", "供应方向", "库存明细", "销售提示"}:
            width = 42
        elif header in {"品名", "异常提示"}:
            width = 28
        elif header in {"SKU", "MSKU", "FNSKU", "ASIN"}:
            width = 18
        sheet.column_dimensions[get_column_letter(column)].width = width


def _write_chart_screenshot_sheet(workbook: Workbook) -> int:
    existing_images = [(title, path) for title, path in CHART_IMAGE_SPECS if path.exists()]
    if not existing_images:
        return 0
    sheet = workbook.create_sheet("图表截图")
    sheet.sheet_view.showGridLines = False
    for column in range(1, 16):
        sheet.column_dimensions[get_column_letter(column)].width = 12
    sheet["A1"] = "SKU 详情图表截图"
    sheet["A1"].font = Font(bold=True, size=15, color="1F1F1F")
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A1:O1")
    current_row = 3
    for title, path in existing_images:
        title_cell = sheet.cell(row=current_row, column=1, value=title)
        title_cell.font = Font(bold=True, size=12, color="1F1F1F")
        title_cell.fill = PatternFill("solid", fgColor="D9EAF7")
        title_cell.alignment = Alignment(vertical="center")
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
        sheet.row_dimensions[current_row].height = 24

        image = ExcelImage(str(path))
        if image.width > CHART_IMAGE_MAX_WIDTH:
            scale = CHART_IMAGE_MAX_WIDTH / image.width
            image.width = int(image.width * scale)
            image.height = int(image.height * scale)
        sheet.add_image(image, f"A{current_row + 1}")
        current_row += max(8, int((image.height + 19) / 20) + 4)
    return len(existing_images)


def _meta_rows(meta: dict[str, Any]) -> list[list[Any]]:
    filters = meta.get("filters") if isinstance(meta.get("filters"), dict) else {}
    return [
        ["导出类型", "当前SKU明细排查"],
        ["销售区间", meta.get("sales_period")],
        ["销售开始日期", meta.get("sales_start_date")],
        ["销售结束日期", meta.get("sales_end_date")],
        ["筛选SKU", meta.get("material_code")],
        ["当前筛选", json.dumps(filters, ensure_ascii=False, default=str)],
        ["命中SKU行数", meta.get("source_rows")],
        ["导出SKU行数", meta.get("loaded_rows")],
        ["最大导出行数", meta.get("max_rows")],
        ["图表截图", meta.get("chart_image_count")],
        ["字段口径", "表头按当前页面详情生成：基础SKU信息 + 断货分段、销售方向、供应方向、库存明细、销售提示；图表截图页嵌入详情页截图。"],
    ]


def _format_risk_summary(item: dict[str, Any], diagnosis: dict[str, Any]) -> str:
    stockout_metrics = diagnosis.get("stockout", {}).get("metrics", {}) if isinstance(diagnosis.get("stockout"), dict) else {}
    summary_parts = []
    if item.get("stockout_risk_level") and item.get("stockout_risk_level") != "normal":
        parts = [
            f"最早第 {_display_value(item.get('pici_first_shortage_days'))} 天",
            f"断 {_display_value(stockout_metrics.get('shortage_days'))} 天",
            f"关键缺口 {_display_value(item.get('pici_key_gap'))}",
        ]
        summary_parts.append("断货：" + "；".join(parts))
    if item.get("overstock_risk_level") and item.get("overstock_risk_level") != "normal":
        summary_parts.append(f"冗余：{item.get('overstock_warning') or _label(RISK_LEVEL_LABELS, item.get('overstock_risk_level')) or '-'}")
    evidence = item.get("evidence") if isinstance(item.get("evidence"), dict) else {}
    risk_flags = evidence.get("risk_flags") if isinstance(evidence.get("risk_flags"), list) else []
    if risk_flags:
        summary_parts.append(f"底表异常：{len(risk_flags)}项")
    return "\n".join(summary_parts)


def _department_root_cause_sections(causes: Any) -> dict[str, dict[str, Any]]:
    rows = [
        cause
        for cause in causes
        if isinstance(cause, dict) and cause.get("type") not in EXPORT_IGNORED_ROOT_CAUSE_TYPES
    ] if isinstance(causes, list) else []
    used: set[int] = set()
    sections: dict[str, dict[str, Any]] = {}
    for definition in DEPARTMENT_SECTION_DEFINITIONS:
        types = definition["types"]
        items: list[dict[str, Any]] = []
        for index, cause in enumerate(rows):
            if index in used or cause.get("type") not in types:
                continue
            used.add(index)
            items.append(cause)
        sections[definition["key"]] = {
            "title": definition["title"],
            "items": items,
            "text": _format_department_cause_items(items),
        }
    other_items = [cause for index, cause in enumerate(rows) if index not in used]
    if other_items:
        sections["inventory"]["items"].extend(other_items)
        sections["inventory"]["text"] = _format_department_cause_items(sections["inventory"]["items"])
    return sections


def _department_section_text(sections: dict[str, Any], key: str) -> str:
    section = sections.get(key) if isinstance(sections.get(key), dict) else {}
    return str(section.get("text") or "")


def _format_sales_signal_text(diagnosis: Any) -> str:
    if not isinstance(diagnosis, dict):
        return ""
    directions = diagnosis.get("direction_recommendations") if isinstance(diagnosis.get("direction_recommendations"), dict) else {}
    sales = directions.get("sales") if isinstance(directions.get("sales"), dict) else {}
    performance = sales.get("sales_performance") if isinstance(sales.get("sales_performance"), dict) else {}
    potential = sales.get("sales_potential") if isinstance(sales.get("sales_potential"), dict) else {}
    forecast = sales.get("forecast_accuracy") if isinstance(sales.get("forecast_accuracy"), dict) else {}
    control = sales.get("stockout_and_sales_control") if isinstance(sales.get("stockout_and_sales_control"), dict) else {}
    rows: list[str] = []
    control_reminder = _clean_control_reminder_text(control.get("reminder"))
    if control_reminder:
        rows.append(f"控销提醒：{control_reminder}")
    for item in [
        *(performance.get("sales_anomalies") if isinstance(performance.get("sales_anomalies"), list) else []),
        *(potential.get("sales_anomalies") if isinstance(potential.get("sales_anomalies"), list) else []),
    ]:
        if not isinstance(item, dict):
            continue
        reason = _clean_reason_text(item.get("reason") or item.get("evidence")) or "-"
        rows.append(f"销量异常：{item.get('label') or item.get('type') or '-'}：{reason}")
    for item in forecast.get("forecast_anomalies") if isinstance(forecast.get("forecast_anomalies"), list) else []:
        if not isinstance(item, dict):
            continue
        reason_value = item.get("reasons")
        if isinstance(reason_value, list):
            reason = "；".join(_clean_reason_text(value) or str(value).strip(" 。") for value in reason_value if value not in {None, ""})
        else:
            reason = _clean_reason_text(item.get("reason") or item.get("evidence"))
        rows.append(f"预估异常：{item.get('label') or item.get('type') or '-'}：{reason or '-'}")
    if rows:
        return "\n".join(_unique_text(rows))
    return _clean_reason_text(sales.get("summary"))


def _format_department_cause_items(items: list[dict[str, Any]]) -> str:
    lines = [_format_department_cause_item(item) for item in items]
    return "\n".join(line for line in lines if line)


def _format_department_cause_item(cause: dict[str, Any]) -> str:
    label = _label(ROOT_CAUSE_TYPE_LABELS, cause.get("type")) or "归因"
    title = _clean_reason_text(cause.get("cause")) or "-"
    evidence = _clean_reason_text(cause.get("evidence")) or "-"
    lines = [f"{label}：{title}"]
    if evidence:
        lines.append(evidence)
    segment = cause.get("segment") if isinstance(cause.get("segment"), dict) else {}
    if segment:
        segment_parts = [
            str(segment.get("label") or "断货段"),
            f"缺口 {segment.get('shortage_quantity') if segment.get('shortage_quantity') not in {None, ''} else '-'} 件",
            f"断 {segment.get('shortage_days') if segment.get('shortage_days') not in {None, ''} else '-'} 天",
        ]
        lines.append("分段：" + "；".join(segment_parts))
        reasons = segment.get("reasons") if isinstance(segment.get("reasons"), list) else []
        for index, reason in enumerate(reason for reason in reasons if isinstance(reason, dict)):
            direction = str(reason.get("direction") or "归因").strip()
            reason_text = _clean_reason_text(reason.get("reason")) or "-"
            detail = _clean_reason_text(reason.get("detail"))
            prefix = chr(ord("a") + index)
            lines.append(f"{prefix}. {direction}：{reason_text}" + (f"；{detail}" if detail else ""))
    return "；".join(lines)


def _format_stockout_reasons(item: dict[str, Any], diagnosis: dict[str, Any], causes: Any) -> str:
    if item.get("stockout_risk_level") in {None, "", "normal"}:
        return ""
    stockout_metrics = diagnosis.get("stockout", {}).get("metrics", {}) if isinstance(diagnosis.get("stockout"), dict) else {}
    lines: list[str] = []
    parts = [
        f"最早第 {_display_value(item.get('pici_first_shortage_days'))} 天",
        f"断 {_display_value(stockout_metrics.get('shortage_days'))} 天",
        f"关键缺口 {_display_value(item.get('pici_key_gap'))}",
    ]
    lines.append("断货缺口：" + "；".join(parts))
    for cause in (causes if isinstance(causes, list) else []):
        if not isinstance(cause, dict) or cause.get("type") not in STOCKOUT_REASON_TYPES:
            continue
        line = _format_reason_cause_item(cause)
        if line:
            lines.append(line)
    return "\n".join(_unique_text(lines)[:ROOT_CAUSE_LIMIT])


def _format_overstock_reasons(item: dict[str, Any], causes: Any) -> str:
    if item.get("overstock_risk_level") in {None, "", "normal"}:
        return ""
    evidence = item.get("evidence") if isinstance(item.get("evidence"), dict) else {}
    lines: list[str] = []
    reason = _clean_reason_text(evidence.get("overstock_reason"))
    if reason:
        return reason
    for cause in (causes if isinstance(causes, list) else []):
        if not isinstance(cause, dict) or cause.get("type") not in OVERSTOCK_REASON_TYPES:
            continue
        line = _format_reason_cause_item(cause)
        if line:
            lines.append(line)
    return "\n".join(_unique_text(lines)[:ROOT_CAUSE_LIMIT])


def _format_reason_cause_item(cause: dict[str, Any]) -> str:
    label = _label(ROOT_CAUSE_TYPE_LABELS, cause.get("type")) or "归因"
    title = _clean_reason_text(cause.get("cause")) or "-"
    segment = cause.get("segment") if isinstance(cause.get("segment"), dict) else {}
    if segment:
        parts = [
            str(segment.get("label") or "断货段"),
            f"第{_display_value(segment.get('start_day'))}-{_display_value(segment.get('end_day'))}天",
            f"缺口 {_display_value(segment.get('shortage_quantity'))} 件",
            f"断 {_display_value(segment.get('shortage_days'))} 天",
        ]
        reasons = segment.get("reasons") if isinstance(segment.get("reasons"), list) else []
        reason_parts = []
        for reason in reasons:
            if not isinstance(reason, dict):
                continue
            direction = str(reason.get("direction") or "归因").strip()
            reason_text = _clean_reason_text(reason.get("reason"))
            detail = _clean_reason_text(reason.get("detail"))
            if reason_text:
                reason_parts.append(f"{direction}：{reason_text}" + (f"；{detail}" if detail else ""))
        if reason_parts:
            parts.append("原因：" + "；".join(reason_parts[:3]))
        return f"{label}：" + "；".join(parts)
    evidence = _clean_reason_text(cause.get("evidence"))
    return f"{label}：{title}" + (f"；{evidence}" if evidence else "")


def _format_root_causes(causes: Any) -> str:
    if not isinstance(causes, list) or not causes:
        return ""
    lines = []
    for cause in causes:
        if not isinstance(cause, dict):
            continue
        label = _label(ROOT_CAUSE_TYPE_LABELS, cause.get("type")) or "归因"
        title = _clean_reason_text(cause.get("cause")) or "-"
        evidence = _clean_reason_text(cause.get("evidence")) or "-"
        lines.append(f"{label}：{title}；证据：{evidence}")
    return "\n".join(lines)


def _format_risk_flags(flags: Any) -> str:
    if not isinstance(flags, list):
        return ""
    lines = []
    for flag in flags:
        if isinstance(flag, dict):
            lines.append("；".join(f"{key}={value}" for key, value in flag.items() if value not in {None, ""}))
        else:
            lines.append(str(flag))
    return "\n".join(line for line in lines if line)


def _join_lines(values: Any) -> str:
    if isinstance(values, list):
        return "\n".join(str(value) for value in values if value not in {None, ""})
    if values is None:
        return ""
    text = str(values)
    return "" if text == "" else text


def _unique_text(values: list[str]) -> list[str]:
    unique = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in unique:
            unique.append(text)
    return unique


def _clean_reason_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text or text == "-":
        return ""
    if "。" in text:
        text = text.split("。", 1)[0]
    text = re.sub(r"[，,；;\s]*(建议|动作|处理逻辑|处理|补救动作)[:：].*$", "", text)
    text = text.replace("需复核", "异常").replace("需要复核", "异常")
    pieces = re.split(r"[；;\n]+", text)
    kept: list[str] = []
    for piece in pieces:
        cleaned = piece.strip(" ，,。")
        if not cleaned:
            continue
        if any(marker in cleaned for marker in ("建议", "动作：", "处理逻辑", "补救动作")):
            continue
        if cleaned.startswith(("动作", "处理", "补救", "执行前", "保持", "优先", "冻结", "复核", "核查", "确认", "催")):
            continue
        kept.append(cleaned)
    return "；".join(kept)


def _clean_control_reminder_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    if not text or text == "-":
        return ""
    text = re.sub(r"[，,；;\s]*(动作|处理逻辑|处理|补救动作)[:：].*$", "", text)
    pieces = re.split(r"[；;\n]+", text)
    kept: list[str] = []
    for piece in pieces:
        cleaned = piece.strip(" ，,。")
        if cleaned:
            kept.append(cleaned)
    return "；".join(kept)


def _display_value(value: Any) -> str:
    return str(value) if value not in {None, ""} else "-"


def _label(labels: dict[str, str], value: Any) -> str:
    text = "" if value is None else str(value)
    return labels.get(text, text)
