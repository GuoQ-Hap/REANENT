from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from pmc_agent.control_tower_recommendations import build_recommendation_export_workbook, build_recommendation_rows
from tests.fake_control_tower import FakeMainRuleConnector


class ControlTowerRecommendationExportTests(unittest.TestCase):
    def test_recommendation_rows_include_control_and_group_dimensions(self) -> None:
        rows, meta = build_recommendation_rows(connector=FakeMainRuleConnector(), max_rows=100)

        self.assertEqual(meta["source_rows"], 2)
        stockout = next(row for row in rows if row["material_code"] == "A100")
        self.assertEqual(stockout["strategy_label"], "当前45天控销口径")
        self.assertEqual(stockout["sales_department"], "North America")
        self.assertEqual(stockout["salesman"], "Alice")
        self.assertIn("控", stockout["control_segments"])
        self.assertGreater(stockout["control_days"], 0)
        self.assertGreater(stockout["control_saved_quantity"], 0)

    def test_flat_stockout_uses_current_flat_control_strategy_only(self) -> None:
        class FlatStockoutConnector(FakeMainRuleConnector):
            def get_inventory_rows(self, query_spec):
                next_rows = []
                for row in super().get_inventory_rows(query_spec):
                    if row["sku"] == "A100":
                        row = row.copy()
                        row["msku_sales_property"] = "平"
                    next_rows.append(row)
                return next_rows

        rows, _ = build_recommendation_rows(connector=FlatStockoutConnector(), max_rows=100)
        a100_rows = [row for row in rows if row["material_code"] == "A100"]
        a100_strategies = {row["strategy_label"] for row in a100_rows}

        self.assertEqual(a100_strategies, {"当前平滞控销口径"})
        self.assertTrue(all(row["urgent_air_quantity"] == 0 and row["standard_air_quantity"] == 0 and row["fast_quantity"] == 0 for row in a100_rows))
        self.assertTrue(all("65" not in row["replenishment_text"] for row in a100_rows))

    def test_flat_control_spreads_late_small_gap_from_recovery_day(self) -> None:
        class LocalGapConnector(FakeMainRuleConnector):
            def get_inventory_rows(self, query_spec):
                rows = []
                for row in super().get_inventory_rows(query_spec):
                    if row["sku"] != "A100":
                        continue
                    row = row.copy()
                    row["msku_sales_property"] = "平"
                    rows.append(row)
                return rows

            def get_pici_shortage_rows(self, store_name=None, table_name=""):
                return [
                    {
                        "fnsku": "X001A100",
                        "store_name": "Amazon US",
                        "chazhi_0_7": "14/39(-25)",
                        "chazhi_0_14": "14/96(-82)",
                        "chazhi_0_21": "98/131(-33)",
                        "chazhi_0_28": "134/168(-34)",
                        "chazhi_0_35": "182/205(-23)",
                        "chazhi_0_42": "182/242(-60)",
                        "chazhi_0_49": "230/286(-56)",
                        "chazhi_0_56": "230/327(-97)",
                        "chazhi_0_63": "266/369(-103)",
                        "chazhi_0_70": "266/411(-145)",
                        "chazhi_0_77": "266/447(-181)",
                        "chazhi_0_84": "266/484(-218)",
                        "chazhi_0_98": "266/548(-282)",
                    }
                ]

        rows, _ = build_recommendation_rows(connector=LocalGapConnector(), max_rows=100)
        stockout = next(row for row in rows if row["material_code"] == "A100")

        self.assertIn("第1-14天 控60%", stockout["control_segments"])
        self.assertIn("第15-56天 控7%", stockout["control_segments"])
        self.assertNotIn("第1-56天 控60%", stockout["control_segments"])
        self.assertNotIn("第48-56天 控28%", stockout["control_segments"])
        self.assertEqual(stockout["slow_quantity"], 73)
        self.assertIn("覆盖61-75天", stockout["replenishment_text"])
        self.assertNotIn("覆盖61天后", stockout["replenishment_text"])

    def test_recommendation_workbook_has_overall_department_person_and_detail_sheets(self) -> None:
        content, meta = build_recommendation_export_workbook(connector=FakeMainRuleConnector(), max_rows=100)
        workbook = load_workbook(BytesIO(content), read_only=True)

        self.assertEqual(workbook.sheetnames, ["整体建议", "部门建议", "个人建议", "SKU明细"])
        self.assertEqual(meta["export_rows"], 2)
        detail_headers = [cell.value for cell in next(workbook["SKU明细"].iter_rows(min_row=1, max_row=1))]
        self.assertIn("控销段", detail_headers)
        self.assertIn("加急空运建议件数", detail_headers)
        self.assertIn("普通空运建议件数", detail_headers)
        self.assertIn("PMC建议", detail_headers)
        self.assertIn("采购建议", detail_headers)
        self.assertIn("销售建议", detail_headers)
        self.assertIn("控销需求缺口", detail_headers)
        self.assertNotIn("控销减少需求", detail_headers)
        self.assertNotIn("补货建议", detail_headers)
        self.assertNotIn("建议动作", detail_headers)
        summary_headers = [cell.value for cell in next(workbook["整体建议"].iter_rows(min_row=1, max_row=1))]
        self.assertIn("PMC建议", summary_headers)
        self.assertIn("采购建议", summary_headers)
        self.assertIn("销售建议", summary_headers)
        self.assertIn("控销需求缺口", summary_headers)
        self.assertNotIn("控销减少需求", summary_headers)
        self.assertNotIn("建议摘要", summary_headers)
        department_values = [
            row[1]
            for row in workbook["部门建议"].iter_rows(min_row=2, values_only=True)
            if row and row[1]
        ]
        self.assertIn("North America", department_values)
        detail_values = list(workbook["SKU明细"].iter_rows(min_row=2, values_only=True))
        b200_strategy = next(row[0] for row in detail_values if row[6] == "B200")
        self.assertEqual(b200_strategy, "当前平滞控销口径")
        a100_detail = next(row for row in detail_values if row[6] == "A100")
        detail_header_index = {header: index for index, header in enumerate(detail_headers)}
        self.assertTrue(a100_detail[detail_header_index["销售建议"]])
        self.assertIsNone(a100_detail[detail_header_index["采购建议"]])


if __name__ == "__main__":
    unittest.main()
