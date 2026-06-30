from __future__ import annotations

from io import BytesIO
import unittest

from openpyxl import load_workbook

from pmc_agent.control_tower_sku_investigation import (
    build_sku_investigation_export_workbook,
    build_sku_investigation_rows,
)
from tests.fake_control_tower import FakeMainRuleConnector


class ControlTowerSkuInvestigationExportTests(unittest.TestCase):
    def test_sku_investigation_rows_follow_current_filters(self) -> None:
        rows, meta = build_sku_investigation_rows(
            filters={"sales_apartment": "North America", "order_by": "risk_then_demand"},
            connector=FakeMainRuleConnector(),
            max_rows=100,
        )

        self.assertEqual(meta["source_rows"], 1)
        self.assertEqual([row["item"]["material_code"] for row in rows], ["A100"])
        self.assertIn("断货归因", rows[0]["attribution"])
        self.assertTrue(rows[0]["root_cause_summary"])
        self.assertIn("超卖", rows[0]["department_sections"]["sales"]["text"])
        self.assertIn("预估异常", rows[0]["department_sections"]["sales"]["text"])
        self.assertEqual(rows[0]["department_sections"]["supply"]["text"], "")
        self.assertIn("断货缺口", rows[0]["stockout_reason_summary"])
        self.assertNotIn("证据：", rows[0]["department_sections"]["supply"]["text"])
        self.assertNotIn("建议", rows[0]["root_cause_summary"])

    def test_sku_investigation_workbook_has_current_detail_headers(self) -> None:
        content, meta = build_sku_investigation_export_workbook(
            connector=FakeMainRuleConnector(),
            max_rows=100,
        )
        workbook = load_workbook(BytesIO(content))

        self.assertEqual(workbook.sheetnames, ["SKU明细排查", "图表截图", "导出说明"])
        self.assertEqual(meta["export_rows"], 2)
        self.assertEqual(meta["chart_image_count"], 2)
        self.assertEqual(len(workbook["图表截图"]._images), 2)
        headers = [cell.value for cell in next(workbook["SKU明细排查"].iter_rows(min_row=1, max_row=1))]
        self.assertEqual(
            [
                "销售部门",
                "销售员",
                "产品经理",
                "店铺",
                "国家",
                "发货国家",
                "账号",
                "SKU",
                "MSKU",
                "FNSKU",
                "ASIN",
                "品名",
                "销售属性",
                "产品属性",
                "季节属性",
                "MSKU状态",
                "MSKU生命周期",
                "风险类型",
                "整体风险等级",
                "风险分",
                "异常提示",
                "断货分段",
                "销售方向",
                "供应方向",
                "库存明细",
                "销售提示",
            ],
            headers,
        )
        for removed_header in (
            "风险摘要",
            "断货原因",
            "断货总天数",
            "冗余原因",
            "FBA长库龄占比",
            "chazhi_0_7",
            "底表异常标记",
            "来源表",
            "处理逻辑",
            "建议动作",
            "证据1",
        ):
            self.assertNotIn(removed_header, headers)
        detail_rows = list(workbook["SKU明细排查"].iter_rows(min_row=2, values_only=True))
        header_index = {header: index for index, header in enumerate(headers)}
        stockout_row = next(row for row in detail_rows if row[header_index["SKU"]] == "A100")
        overstock_row = next(row for row in detail_rows if row[header_index["SKU"]] == "B200")

        self.assertIsNone(stockout_row[header_index["断货分段"]])
        self.assertIn("超卖", stockout_row[header_index["销售方向"]])
        self.assertIn("预估异常", stockout_row[header_index["销售方向"]])
        self.assertIsNone(stockout_row[header_index["供应方向"]])
        self.assertIsNone(stockout_row[header_index["库存明细"]])
        self.assertIn("控销提醒", stockout_row[header_index["销售提示"]])
        self.assertIn("旺款命中断货风险", stockout_row[header_index["销售提示"]])
        self.assertIn("控销时间：第1-7天 控50%", stockout_row[header_index["销售提示"]])
        self.assertIn("控销量：合计控 40 件", stockout_row[header_index["销售提示"]])
        self.assertIn("销量异常", stockout_row[header_index["销售提示"]])
        self.assertNotIn("2026W", stockout_row[header_index["销售提示"]])
        self.assertNotIn("实际/折算销量", stockout_row[header_index["销售提示"]])
        self.assertNotIn("未命中预估异常", stockout_row[header_index["销售提示"]])
        self.assertIn("较1个月之前的预估线", stockout_row[header_index["销售提示"]])
        self.assertIsNone(overstock_row[header_index["断货分段"]])
        self.assertIsNone(overstock_row[header_index["库存明细"]])
        self.assertIn("计划异常", overstock_row[header_index["供应方向"]])
        self.assertIn("销售方向", overstock_row[header_index["销售提示"]])
        for row in detail_rows:
            for value in row:
                text = str(value or "")
                self.assertNotIn("证据：", text)
                self.assertNotIn("证据1", text)
                self.assertNotIn("fnsku_out_of_stock_risk", text)
                self.assertNotIn("建议", text)
                self.assertNotIn("动作：", text)
                self.assertNotIn("处理逻辑", text)


if __name__ == "__main__":
    unittest.main()
